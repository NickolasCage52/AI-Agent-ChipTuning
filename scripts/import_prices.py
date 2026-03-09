#!/usr/bin/env python3
"""
Импорт файлов прайса в локальную БД (SQLite).

Использование:
  python -m scripts.import_prices
  python -m scripts.import_prices --base data/price_sources/base.xlsx --defect data/price_sources/defect.xlsx
  python -m scripts.import_prices --base data/price_sources/base.csv --defect data/price_sources/defect.csv
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sqlite3
import sys
from pathlib import Path
from zipfile import BadZipFile

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

ROOT = Path(__file__).parent.parent
DB_PATH = os.getenv("DB_PATH", str(ROOT / "data" / "parts.db"))

# Маппинг: имя колонки в файле -> поле в БД (для обоих прайсов)
# Реальные заголовки из data/price_sources/base.xlsx:
# Номенклатура, Бренд, Артикул, Описание, Кратность отгрузки, Цена руб., Наличие, Срок поставки дн., Каталожный номер, OEМ Номер
COLUMN_ALIASES: dict[str, str] = {
    "Номенклатура": "nomenclature",
    "Наименование": "nomenclature",
    "Бренд": "brand",
    "Артикул": "article_raw",
    "Описание": "description",
    "Кратность отгрузки": "batch_size",
    "Цена": "price",
    "Цена, руб.": "price",
    "Цена руб.": "price",
    "Цена (руб.)": "price",
    "Наличие": "in_stock",
    "Срок поставки (дн.)": "delivery_days",
    "Срок поставки, дн.": "delivery_days",
    "Срок поставки дн.": "delivery_days",
    "Срок": "delivery_days",
    "Срок (дн.)": "delivery_days",
    "Каталожный номер": "catalog_number",
    "OEM Номер": "oem_number",
    "OEМ Номер": "oem_number",  # Cyrillic М
    "OEM": "oem_number",
    "Вес/Объем": "weight_volume",
    "Применимость": "applicability",
}


def normalize_article(raw: str) -> str:
    if not raw:
        return ""
    return re.sub(r"[\s\-_]", "", str(raw)).upper()


def _get_column_map(file_headers: list[str]) -> dict[str, str]:
    """Вернуть маппинг: заголовок_в_файле -> поле_в_БД."""
    result = {}
    for h in file_headers:
        h_clean = (h or "").strip()
        if not h_clean:
            continue
        if h_clean in COLUMN_ALIASES:
            result[h_clean] = COLUMN_ALIASES[h_clean]
    return result


def _detect_format_and_path(filepath: str) -> tuple[str, str]:
    """
    Определить формат по содержимому.
    Возвращает (ext, path): если .csv содержит XLSX, вернёт временный .xlsx путь.
    """
    path = Path(filepath)
    try:
        with open(filepath, "rb") as f:
            head = f.read(4)
        if head[:2] == b"PK" and path.suffix.lower() == ".csv":
            import tempfile
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            with open(filepath, "rb") as src:
                tmp.write(src.read())
            tmp.close()
            return ".xlsx", tmp.name
    except Exception:
        pass
    return path.suffix.lower(), filepath


def read_file(filepath: str, expected_columns: dict | None = None) -> tuple[list[dict], list[str]]:
    """Читает XLSX или CSV, возвращает (rows, errors)."""
    path = Path(filepath)
    errors = []

    if not path.exists():
        return [], [f"Файл не найден: {filepath}"]

    ext, read_path = _detect_format_and_path(filepath)
    _cleanup_path = read_path if read_path != filepath else None
    raw_rows = None

    try:
        if ext in (".xlsx", ".xls"):
            xlsx_ok = False
            if HAS_OPENPYXL:
                try:
                    wb = openpyxl.load_workbook(read_path, read_only=True, data_only=True)
                    ws = wb.active
                    headers = [
                        str(cell.value).strip() if cell.value else ""
                        for cell in next(ws.iter_rows(max_row=1))
                    ]
                    raw_rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        raw_rows.append(
                            dict(zip(headers, [str(v) if v is not None else "" for v in row]))
                        )
                    wb.close()
                    xlsx_ok = True
                except BadZipFile:
                    errors.append(f"Файл {path.name} не является XLSX. Пробуем xlrd (xls)...")
            if not xlsx_ok and HAS_PANDAS:
                try:
                    df = pd.read_excel(read_path, dtype=str, engine="openpyxl")
                    raw_rows = df.to_dict("records")
                    xlsx_ok = True
                except (BadZipFile, Exception):
                    try:
                        import xlrd
                        df = pd.read_excel(read_path, dtype=str, engine="xlrd")
                        raw_rows = df.to_dict("records")
                        xlsx_ok = True
                    except (ImportError, Exception):
                        pass
            if not xlsx_ok:
                ext = ".csv"
                read_path = filepath
        if ext == ".csv":
            encodings = ["utf-8-sig", "utf-8", "cp1251", "cp866", "cp1252", "latin-1", "iso-8859-1"]
            raw_rows = None
            for enc in encodings:
                try:
                    with open(read_path, encoding=enc, newline="", errors="strict") as f:
                        sample = f.read(8192)
                        f.seek(0)
                        try:
                            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                        except csv.Error:
                            dialect = csv.excel()
                        reader = csv.DictReader(f, dialect=dialect, restkey="_extra", restval="")
                        raw_rows = [dict(r) for r in reader]
                    break
                except (UnicodeDecodeError, Exception):
                    continue
            if raw_rows is None:
                hint = "Попробуйте сохранить CSV в UTF-8: Excel → Сохранить как → CSV UTF-8"
                return [], [f"Не удалось прочитать CSV: {path.name}. {hint}"]
        elif ext not in (".xlsx", ".xls", ".csv"):
            return [], [f"Неподдерживаемый формат: {ext}. Используйте XLSX или CSV."]
    finally:
        if _cleanup_path and os.path.exists(_cleanup_path):
            try:
                os.unlink(_cleanup_path)
            except Exception:
                pass

    if not raw_rows:
        return [], [f"Файл пустой: {filepath}"]

    file_headers = list(raw_rows[0].keys())
    col_map = _get_column_map(file_headers)
    if not col_map:
        errors.append(
            f"⚠️ В файле {path.name} отсутствуют нужные колонки. "
            f"Найдены: {file_headers}. Поддерживаются: {list(COLUMN_ALIASES.keys())}"
        )

    rows: list[dict[str, object]] = []
    for i, raw in enumerate(raw_rows):
        try:
            row: dict[str, object] = {}
            for src_col, dst_col in col_map.items():
                val = raw.get(src_col, "")
                row[dst_col] = (
                    str(val).strip()
                    if val and str(val).strip() not in ("None", "nan", "")
                    else None
                )
            row["article"] = normalize_article(row.get("article_raw") or "")
            if row.get("nomenclature") and not row.get("description"):
                row["description"] = row["nomenclature"]
            if row.get("description") and not row.get("nomenclature"):
                row["nomenclature"] = row["description"]
            for field in ("price", "batch_size"):
                if row.get(field):
                    try:
                        row[field] = float(
                            str(row[field]).replace(",", ".").replace(" ", "")
                        )
                    except ValueError:
                        row[field] = None
            if row.get("delivery_days") is not None:
                try:
                    row["delivery_days"] = int(
                        float(str(row["delivery_days"]).replace(",", "."))
                    )
                except ValueError:
                    row["delivery_days"] = None
            rows.append(row)
        except Exception as e:
            errors.append(f"Строка {i+2}: {e}")

    return rows, errors


def get_table_columns(table: str) -> list[str]:
    if table == "products_defect":
        return [
            "nomenclature", "brand", "article", "article_raw", "description",
            "weight_volume", "batch_size", "price", "in_stock", "delivery_days",
            "catalog_number", "oem_number", "applicability",
            "source_file", "import_run_id",
        ]
    return [
        "nomenclature", "brand", "article", "article_raw", "description",
        "batch_size", "price", "in_stock", "delivery_days",
        "catalog_number", "oem_number", "source_file", "import_run_id",
    ]


def import_file(
    conn: sqlite3.Connection,
    filepath: str,
    table: str,
    columns: dict,
    file_type: str,
) -> None:
    rows, errors = read_file(filepath, columns)

    for e in errors:
        print(f"  ⚠️ {e}")

    if not rows:
        print(f"  ❌ Нет данных из {filepath}")
        conn.execute(
            "INSERT INTO import_runs (file_type, filename, rows_imported, rows_failed, errors_json) VALUES (?,?,?,?,?)",
            (file_type, Path(filepath).name, 0, 0, json.dumps(errors, ensure_ascii=False)),
        )
        conn.commit()
        return

    conn.execute(f"DELETE FROM {table}")
    run_id = conn.execute(
        "INSERT INTO import_runs (file_type, filename, rows_imported, rows_failed, errors_json) VALUES (?,?,?,?,?)",
        (file_type, Path(filepath).name, 0, 0, "[]"),
    ).lastrowid

    allowed_cols = get_table_columns(table)
    imported = 0
    failed = 0
    row_errors = []

    for i, row in enumerate(rows):
        row["source_file"] = Path(filepath).name
        row["import_run_id"] = run_id
        cols = [c for c in row.keys() if c in allowed_cols]
        try:
            placeholders = ", ".join(["?"] * len(cols))
            col_names = ", ".join(cols)
            values = [row.get(c) for c in cols]
            conn.execute(
                f"INSERT INTO {table} ({col_names}) VALUES ({placeholders})", values
            )
            imported += 1
        except Exception as e:
            failed += 1
            row_errors.append(f"Строка {i+1}: {e}")

    conn.execute(
        "UPDATE import_runs SET rows_imported=?, rows_failed=?, errors_json=? WHERE id=?",
        (imported, failed, json.dumps(row_errors[:20], ensure_ascii=False), run_id),
    )
    conn.commit()
    print(f"  ✅ Импортировано: {imported} строк | Ошибок: {failed}")


def main() -> None:
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass
    parser = argparse.ArgumentParser(description="Импорт прайсов в БД")
    base_default = ROOT / "data" / "price_sources" / "base.csv"
    if not base_default.exists():
        base_default = ROOT / "data" / "price_sources" / "base.xlsx"
    defect_default = ROOT / "data" / "price_sources" / "defect.csv"
    if not defect_default.exists():
        defect_default = ROOT / "data" / "price_sources" / "defect.xlsx"
    parser.add_argument("--base", default=str(base_default), help="Путь к прайсу 1 (базовый)")
    parser.add_argument("--defect", default=str(defect_default), help="Путь к прайсу 2 (некондиция)")
    parser.add_argument("--db", default=DB_PATH, help="Путь к SQLite БД")
    args = parser.parse_args()

    print(f"\n📦 Импорт прайсов в БД: {args.db}\n")

    Path(args.db).parent.mkdir(parents=True, exist_ok=True)
    (ROOT / "data" / "price_sources").mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row

    conn.executescript("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nomenclature TEXT, brand TEXT, article TEXT, article_raw TEXT,
            description TEXT, batch_size REAL, price REAL, in_stock TEXT,
            delivery_days INTEGER, catalog_number TEXT, oem_number TEXT,
            source_file TEXT, import_run_id INTEGER
        );
        CREATE TABLE IF NOT EXISTS products_defect (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nomenclature TEXT, brand TEXT, article TEXT, article_raw TEXT,
            description TEXT, weight_volume TEXT, batch_size REAL, price REAL,
            in_stock TEXT, delivery_days INTEGER, catalog_number TEXT,
            oem_number TEXT, applicability TEXT, source_file TEXT, import_run_id INTEGER
        );
        CREATE TABLE IF NOT EXISTS import_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_type TEXT, filename TEXT, rows_imported INTEGER,
            rows_failed INTEGER, errors_json TEXT,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_products_article ON products(article);
        CREATE INDEX IF NOT EXISTS idx_products_oem ON products(oem_number);
        CREATE INDEX IF NOT EXISTS idx_products_catalog ON products(catalog_number);
        CREATE INDEX IF NOT EXISTS idx_defect_article ON products_defect(article);
        CREATE INDEX IF NOT EXISTS idx_defect_oem ON products_defect(oem_number);
    """)

    print("1️⃣  Прайс базовый:")
    import_file(conn, args.base, "products", COLUMN_ALIASES, "base")

    print("2️⃣  Прайс некондиция:")
    import_file(conn, args.defect, "products_defect", COLUMN_ALIASES, "defect")

    conn.close()
    print("\n✅ Готово. Данные загружены в БД.\n")


if __name__ == "__main__":
    main()

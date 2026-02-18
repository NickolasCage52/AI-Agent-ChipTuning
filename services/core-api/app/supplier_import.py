from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook


@dataclass
class NormalizedOffer:
    sku: str | None
    oem: str | None
    name: str | None
    brand: str | None
    price: Decimal | None
    stock: int | None
    delivery_days: int | None


_HEADER_ALIASES: dict[str, str] = {
    "sku": "sku",
    "артикул": "sku",
    "article": "sku",
    "partnumber": "sku",
    "oem": "oem",
    "oe": "oem",
    "оригинал": "oem",
    "номер": "oem",
    "name": "name",
    "наименование": "name",
    "товар": "name",
    "brand": "brand",
    "бренд": "brand",
    "manufacturer": "brand",
    "price": "price",
    "цена": "price",
    "стоимость": "price",
    "stock": "stock",
    "остаток": "stock",
    "наличие": "stock",
    "qty": "stock",
    "delivery_days": "delivery_days",
    "срок": "delivery_days",
    "доставка": "delivery_days",
}


def _norm_header(h: Any) -> str:
    return str(h or "").strip().lower().replace(" ", "_").replace("-", "_")


def _to_decimal(v: Any) -> Decimal | None:
    s = str(v or "").strip()
    if not s:
        return None
    s = s.replace(" ", "").replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return None


def _to_int(v: Any) -> int | None:
    s = str(v or "").strip()
    if not s:
        return None
    s = s.replace(" ", "")
    try:
        return int(float(s))
    except Exception:
        return None


def _map_headers(raw_headers: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for idx, h in enumerate(raw_headers):
        nh = _norm_header(h)
        key = _HEADER_ALIASES.get(nh)
        if key:
            mapping[idx] = key
    return mapping


def parse_supplier_price(filename: str, content: bytes) -> list[NormalizedOffer]:
    fn = (filename or "").lower()
    if fn.endswith(".xlsx") or fn.endswith(".xlsm") or fn.endswith(".xltx"):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> list[NormalizedOffer]:
    text = content.decode("utf-8", errors="ignore")
    # try delimiter autodetect
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t")
    except Exception:
        dialect = csv.get_dialect("excel")
        dialect.delimiter = ";"  # type: ignore[attr-defined]
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows = list(reader)
    if not rows:
        return []
    header = rows[0]
    mapping = _map_headers(header)
    offers: list[NormalizedOffer] = []
    for r in rows[1:]:
        offers.append(_row_to_offer(r, mapping))
    return [o for o in offers if (o.sku or o.oem or o.name)]


def _parse_xlsx(content: bytes) -> list[NormalizedOffer]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return []
    mapping = _map_headers(header)
    offers: list[NormalizedOffer] = []
    for r in rows_iter:
        offers.append(_row_to_offer(list(r), mapping))
    return [o for o in offers if (o.sku or o.oem or o.name)]


def _row_to_offer(row: list[Any], mapping: dict[int, str]) -> NormalizedOffer:
    data: dict[str, Any] = {}
    for idx, key in mapping.items():
        if idx < len(row):
            data[key] = row[idx]
    return NormalizedOffer(
        sku=(str(data.get("sku")).strip() if data.get("sku") not in (None, "") else None),
        oem=(str(data.get("oem")).strip() if data.get("oem") not in (None, "") else None),
        name=(str(data.get("name")).strip() if data.get("name") not in (None, "") else None),
        brand=(str(data.get("brand")).strip() if data.get("brand") not in (None, "") else None),
        price=_to_decimal(data.get("price")),
        stock=_to_int(data.get("stock")),
        delivery_days=_to_int(data.get("delivery_days")),
    )

